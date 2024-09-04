import cv2
import numpy as np
import face_recognition
from datetime import datetime
from openpyxl import load_workbook
import os

# Function to load training images and their names
def load_training_data(path):
  """Loads images from a directory and extracts names as class names.

  Args:
      path: Path to the directory containing training images.

  Returns:
      A tuple containing two lists:
          - images: List of loaded training images.
          - classNames: List of class names corresponding to the images.
  """
  images = []
  classNames = []
  for filename in os.listdir("C:\\GCU\\Pyhton programs\\Project\\Face_recognition_attendance\\3rd_attendance(Main)\\known_faces"): #"""path of the folder which contain training image"""
    img = cv2.imread(os.path.join(path, filename))
    if img is not None:
      classNames.append(os.path.splitext(filename)[0])
      images.append(img)
  return images, classNames

# Function to encode faces in training images
def find_encodings(images):
  """Encodes faces in a list of images.

  Args:
      images: List of training images.

  Returns:
      A list of face encodings corresponding to the input images.
  """
  encodeList = []
  for img in images:
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    encode = face_recognition.face_encodings(img)[0]
    encodeList.append(encode)
  return encodeList

# Function to mark attendance in a spreadsheet
def mark_attendance(roll_no, date):
  """Marks attendance for a roll number on a specific date in a spreadsheet.

  Args:
      roll_no: Student's roll number.
      date: Date of attendance (YYYY-MM-DD format).
  """
  try:
    # Open the Excel workbook
    wb = load_workbook("C:\\GCU\\Pyhton programs\\Project\\Face_recognition_attendance\\3rd_attendance(Main)\\lolattendance_list.xlsx")
    sheet = wb.active

    # Find or create the date column
    date_column = None
    last_column = None
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=3):
      if col[0].value is not None:
        last_column = col[0].column
        if col[0].value == date:
          date_column = col[0].column
          break
    if date_column is None:
      if last_column is None:
        date_column = 3
      else:
        date_column = last_column + 1
      sheet.cell(row=1, column=date_column, value=date)

    # Find or create the student row
    rollno_row = None
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
      if str(row[0].value) == roll_no:
        rollno_row = row[0].row
        break
    if rollno_row is None:
      rollno_row = sheet.max_row + 1
      sheet.cell(row=rollno_row, column=1, value=roll_no)

    # Mark attendance (if not already marked)
    if sheet.cell(row=rollno_row, column=date_column).value != "P":
      sheet.cell(row=rollno_row, column=date_column, value="P")
      print(f'Attendance marked for {roll_no}')
    else:
      print(f'Attendance already marked for {roll_no} on {date}')

    # Save changes to the spreadsheet
    wb.save("C:\\GCU\\Pyhton programs\\Project\\Face_recognition_attendance\\3rd_attendance(Main)\\lolattendance_list.xlsx")

  except Exception as e:
    print(f'Error occurred while marking attendance: {e}')

def main():
  """Main function that loads training data, performs face recognition, and marks attendance."""

  # Path to the training images directory
  path = "C:\\GCU\\Pyhton programs\\Project\\Face_recognition_attendance\\3rd_attendance(Main)\\known_faces"

  # Load training images and names
  images, classNames = load_training_data(path)

  # Encode known faces
  encodeListKnown = find_encodings(images)
  print('Encoding Complete')

  # Start video capture
  cap = cv2.VideoCapture(0)

  recognized_faces = []
  while True:
        # Capture a frame from the webcam
    success, img = cap.read()

    # Resize the frame for faster processing
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    # Find faces in the current frame
    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    # Loop through each detected face
    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
      matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
      faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

      # Find the closest match (if any)
      matchIndex = np.argmin(faceDis)

      # Check if there's a match
      if matches[matchIndex]:
        rollno = classNames[matchIndex].upper()
        name = classNames[matchIndex]

        # Check if face has already been recognized to avoid duplicates
        if rollno not in recognized_faces:
          recognized_faces.append(rollno)

          # Extract face location coordinates and scale them back to original frame size
          y1, x2, y2, x1 = faceLoc
          y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4

          # Draw a rectangle around the recognized face
          cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
          cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
          cv2.putText(img, rollno, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

          # Mark attendance for the recognized student
          mark_attendance(rollno, str(datetime.now().strftime("%Y-%m-%d")))

    # Exit loop on 'q' key press
    if cv2.waitKey(1) & 0xFF == ord('q'):
      break

    # Display the webcam frame
    cv2.imshow('Webcam', img)
    cv2.waitKey(1)

  # Release resources
  cap.release()
  cv2.destroyAllWindows()

if __name__ == "__main__":
  main()
