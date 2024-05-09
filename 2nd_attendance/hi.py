
#importing the modules/packages

import cv2
import numpy as np
import face_recognition
from datetime import datetime, date
from openpyxl import load_workbook
import os

# Function to load training images and their corresponding class names
def loadTrainingData(path):
    images = [cv2.imread(f'{path}/{cl}') for cl in os.listdir(path)]
    classNames = [os.path.splitext(cl)[0] for cl in os.listdir(path)]
    return images, classNames

# Function to encode faces in the training images
def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList

#function with two paraameters
def markAttendance(image_rollno, date):
    try:
        # Current date
        now = datetime.now()
        date_format = now.strftime("%Y-%m-%d")

        # Opening the Excel file
        wb = load_workbook("C:\\GCU\\Pyhton programs\\Project\\PY\\lol.xlsx")
        sheet = wb.active

        # Find the column index for today's date or create it if not present
        date_column = None
        last_column = None
        for col in sheet.iter_cols(min_row=1, max_row=1, min_col=3):
            if col[0].value is not None:
                last_column = col[0].column
            if col[0].value == date_format:
                date_column = col[0].column
                break
        if date_column is None:
            if last_column is None:
                date_column = 3  # If no data exists, place the date column in column C
            else:
                date_column = last_column + 1  # Place the date column after the last column containing data
            sheet.cell(row=1, column=date_column, value=date_format)

        # Find the row for the roll number or create it if not present
        rollno_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            if str(row[0].value) == image_rollno:
                rollno_row = row[0].row
                break
        if rollno_row is None:
            rollno_row = sheet.max_row + 1
            sheet.cell(row=rollno_row, column=1, value=image_rollno)

        # Mark attendance
        if sheet.cell(row=rollno_row, column=date_column).value != "P":
            sheet.cell(row=rollno_row, column=date_column, value="P")
            print(f'Attendance marked for {image_rollno}')
        else:
            print(f'Attendance already marked for {image_rollno} on {date_format}')

        # Save changes
        wb.save('lol.xlsx')

    except Exception as e:
        print(f'Error occurred while marking attendance: {e}')


recognized_faces = []

def main():
    #path to the file
    path = "C:\\GCU\\Pyhton programs\\Project\\PY\\Training_images"

    #loading training images and their corresponding class names
    images, classNames = loadTrainingData(path)

    #encoding known faces
    encodeListKnown = findEncodings(images)
    print('Encoding Complete')

    #initializing teh video capture
    cap = cv2.VideoCapture(0)

    while True:
        success, img = cap.read()
        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                rollno = classNames[matchIndex].upper()
                name = classNames[matchIndex]

                # Check if face has already been recognized
                if rollno not in recognized_faces:
                    recognized_faces.append(rollno)

                    y1, x2, y2, x1 = faceLoc
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                    cv2.putText(img, rollno, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    markAttendance(rollno, name)
                    break

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

        cv2.imshow('Webcam', img)
        cv2.waitKey(1)

if __name__ == "__main__":
    main()
